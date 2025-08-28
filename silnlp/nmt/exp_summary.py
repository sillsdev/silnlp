import argparse
import glob
import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import get_mt_exp_dir

chap_num = 0


def read_group_results(
    file_path: str,
    target_book: str,
    all_books: list[str],
    metrics: list[str],
    key_word: str,
) -> tuple[dict[str, dict[int, list[str]]], set[int]]:
    global chap_num

    data = {}
    chapter_groups = set()
    for lang_pair in os.listdir(file_path):
        lang_pattern = re.compile(r"([\w-]+)\-([\w-]+)")
        if not lang_pattern.match(lang_pair):
            continue

        data[lang_pair] = {}
        prefix = "+".join(all_books)
        pattern = re.compile(rf"^{re.escape(prefix)}_{key_word}_order_(\d+)_ch$")

        for groups in os.listdir(os.path.join(file_path, lang_pair)):
            if m := pattern.match(os.path.basename(groups)):
                folder_path = os.path.join(file_path, lang_pair, os.path.basename(groups))
                diff_pred_file = glob.glob(os.path.join(folder_path, "diff_predictions*"))
                if diff_pred_file:
                    r = extract_diff_pred_data(diff_pred_file[0], metrics, target_book)
                    data[lang_pair][int(m.group(1))] = r
                else:
                    data[lang_pair][int(m.group(1))] = {}
                    print(folder_path + " has no diff_predictions file.")
                chapter_groups.add(int(m.group(1)))
                chap_num = max(chap_num, int(m.group(1)))
    return data, chapter_groups


def extract_diff_pred_data(filename: str, metrics: list[str], target_book: str, header_row=5) -> dict[int, list[str]]:
    global chap_num

    metrics = [m.lower() for m in metrics]
    try:
        df = pd.read_excel(filename, header=header_row)
    except ValueError as e:
        print(f"An error occurs in {filename}")
        print(e)
        return {}

    df.columns = [col.strip().lower() for col in df.columns]

    result = {}
    metric_warning = False
    for _, row in df.iterrows():
        vref = row["vref"]
        m = re.match(r"(\d?[A-Z]{2,3}) (\d+)", str(vref))
        if not m:
            print(f"Invalid VREF format: {str(vref)}")
            return {}

        book_name, chap = m.groups()
        if book_name != target_book:
            continue

        chap_num = max(chap_num, int(chap))
        values = []
        for metric in metrics:
            if metric in row:
                values.append(row[metric])
            else:
                metric_warning = True
                values.append(None)

        result[int(chap)] = values

    if metric_warning:
        print("Warning: {metric} was not calculated in {filename}")

    return result


def flatten_dict(data: dict, chapter_groups: list[int], metrics: list[str], baseline={}) -> list[str]:
    global chap_num

    rows = []
    if len(data) > 0:
        for lang_pair in data:
            for chap in range(1, chap_num + 1):
                row = [lang_pair, chap]
                row.extend([None, None, None] * len(metrics) * len(data[lang_pair]))
                row.extend([None] * len(chapter_groups))
                row.extend([None] * (1 + len(metrics)))

                for res_chap in data[lang_pair]:
                    if chap in data[lang_pair][res_chap]:
                        for m in range(len(metrics)):
                            index_m = (
                                3 + 1 + len(metrics) + chapter_groups.index(res_chap) * (len(metrics) * 3 + 1) + m * 3
                            )
                            row[index_m] = data[lang_pair][res_chap][chap][m]
                if len(baseline) > 0:
                    for m in range(len(metrics)):
                        row[3 + m] = baseline[lang_pair][chap][m] if lang_pair in baseline else None
                rows.append(row)
    else:
        for lang_pair in baseline:
            for chap in range(1, chap_num + 1):
                row = [lang_pair, chap]
                row.extend([None] * (1 + len(metrics)))

                for m in range(len(metrics)):
                    row[3 + m] = baseline[lang_pair][chap][m]
                rows.append(row)

    return rows


def create_xlsx(rows: list[str], chapter_groups: list[str], output_path: str, metrics: list[str]) -> None:
    global chap_num

    wb = Workbook()
    ws = wb.active

    num_col = len(metrics) * 3 + 1
    groups = [("language pair", 1), ("Chapter", 1), ("Baseline", (1 + len(metrics)))]
    for chap in chapter_groups:
        groups.append((chap, num_col))

    col = 1
    for header, span in groups:
        start_col = get_column_letter(col)
        end_col = get_column_letter(col + span - 1)
        ws.merge_cells(f"{start_col}1:{end_col}1")
        ws.cell(row=1, column=col, value=header)
        col += span

    sub_headers = []
    baseline_headers = []

    for i, metric in enumerate(metrics):
        if i == 0:
            baseline_headers.append("rank")
            sub_headers.append("rank")
        baseline_headers.append(metric)
        sub_headers.append(metric)
        sub_headers.append("diff (prev)")
        sub_headers.append("diff (start)")

    for i, baseline_header in enumerate(baseline_headers):
        ws.cell(row=2, column=3 + i, value=baseline_header)

    col = 3 + len(metrics) + 1
    for _ in range(len(groups) - 3):
        for i, sub_header in enumerate(sub_headers):
            ws.cell(row=2, column=col + i, value=sub_header)

        col += len(sub_headers)

    for row in rows:
        ws.append(row)

    for row_idx in [1, 2]:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    cur_lang_pair = 3
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row=row_idx, column=4).value is not None:
            ws.cell(row=row_idx, column=3).value = (
                f"=RANK.EQ(D{row_idx}, INDEX(D:D, INT((ROW(D{row_idx})-3)/{chap_num})*{chap_num}+3):INDEX(D:D, \
                        INT((ROW(D{row_idx})-3)/{chap_num})*{chap_num}+{chap_num}+2), 0)"
            )

        start_col = 3 + len(metrics) + 1
        end_col = ws.max_column

        while start_col < end_col:
            start_col += 1
            if ws.cell(row=row_idx, column=start_col).value is None:
                for col in range(start_col - 1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        fill_type="solid", start_color="CCCCCC", end_color="CCCCCC"
                    )
                break

            col_letter = get_column_letter(start_col)
            ws.cell(row=row_idx, column=start_col - 1).value = (
                f"=RANK.EQ({col_letter}{row_idx}, INDEX({col_letter}:{col_letter}, \
                    INT((ROW({col_letter}{row_idx})-3)/{chap_num})*{chap_num}+3):INDEX({col_letter}:{col_letter}, \
                        INT((ROW({col_letter}{row_idx})-3)/{chap_num})*{chap_num}+{chap_num}+2), 0)"
            )

            for i in range(1, len(metrics) + 1):
                start_letter = get_column_letter(3 + i)

                diff_prev_col = start_col + 1
                diff_start_col = start_col + 2

                prev_letter = (
                    start_letter
                    if diff_prev_col <= 3 + len(metrics) + 1 + 3 * len(metrics)
                    else get_column_letter(diff_prev_col - 1 - 1 - 3 * len(metrics))
                )
                cur_letter = get_column_letter(diff_prev_col - 1)

                ws.cell(row=row_idx, column=diff_prev_col).value = f"={cur_letter}{row_idx}-{prev_letter}{row_idx}"
                ws.cell(row=row_idx, column=diff_start_col).value = f"={cur_letter}{row_idx}-{start_letter}{row_idx}"

                start_col += 3

        if ws.cell(row=row_idx, column=1).value != ws.cell(row=cur_lang_pair, column=1).value:
            ws.merge_cells(start_row=cur_lang_pair, start_column=1, end_row=row_idx - 1, end_column=1)
            cur_lang_pair = row_idx
        elif row_idx == ws.max_row:
            ws.merge_cells(start_row=cur_lang_pair, start_column=1, end_row=row_idx, end_column=1)

    wb.save(output_path)


# Sample command:
# python -m silnlp.nmt.exp_summary Catapult_Reloaded_Confidences
# --trained-books MRK --target-book MAT --metrics chrf3 confidence --key-word conf --baseline Catapult_Reloaded/2nd_book/MRK
def main() -> None:
    global chap_num

    parser = argparse.ArgumentParser(
        description="Pulling results from a single experiment and/or multiple experiment groups."
        "A valid experiment should have the following format:"
        "baseline/lang_pair/exp_group/diff_predictions or baseline/lang_pair/diff_predictions for a single experiment"
        "or "
        "exp/lang_pair/exp_groups/diff_predictions for multiple experiment groups"
        "More information in --exp and --baseline."
        "Use --exp for multiple experiment groups and --baseline for a single experiment."
        "At least one --exp or --baseline needs to be specified."
    )
    parser.add_argument(
        "--exp",
        type=str,
        help="Experiment folder with progression results. "
        "A valid experiment groups should have the following format:"
        "exp/lang_pair/exp_groups/diff_predictions"
        "where there should be at least one exp_groups that naming in the following format:"
        "*book*+*book*_*key-word*_order_*number*_ch"
        "where *book*+*book*... are the combination of all --trained-books with the last one being --target-book."
        "More information in --key-word.",
    )
    parser.add_argument(
        "--trained-books", nargs="*", required=True, type=str.upper, help="Books that are trained in the exp"
    )
    parser.add_argument("--target-book", required=True, type=str.upper, help="Book that is going to be analyzed")
    parser.add_argument(
        "--metrics",
        nargs="*",
        metavar="metrics",
        default=["chrf3", "confidence"],
        type=str.lower,
        help="Metrics that will be analyzed with",
    )
    parser.add_argument(
        "--key-word",
        type=str,
        default="conf",
        help="Key word in the filename for the exp group to distinguish between the experiment purpose."
        "For example, in LUK+ACT_conf_order_12_ch, the key-word should be conf."
        "Another example, in LUK+ACT_standard_order_12_ch, the key-word should be standard.",
    )
    parser.add_argument(
        "--baseline",
        type=str,
        help="A non-progression folder for a single experiment."
        "A valid single experiment should have the following format:"
        "baseline/lang_pair/exp_group/diff_predictions where exp_group will be in the following format:"
        "*book*+*book*... as the combination of all --trained-books."
        "or"
        "baseline/lang_pair/diff_predictions "
        "where the information of --trained-books should have already been indicated in the baseline name.",
    )
    args = parser.parse_args()

    if not (args.exp or args.baseline):
        parser.error("At least one --exp or --baseline needs to be specified.")

    trained_books = args.trained_books
    target_book = args.target_book
    all_books = trained_books + [target_book]
    metrics = args.metrics
    key_word = args.key_word

    multi_group_exp_name = args.exp
    multi_group_exp_dir = get_mt_exp_dir(multi_group_exp_name) if multi_group_exp_name else None

    single_group_exp_name = args.baseline
    single_group_exp_dir = get_mt_exp_dir(single_group_exp_name) if single_group_exp_name else None

    result_file_name = "+".join(all_books)
    result_dir = multi_group_exp_dir if multi_group_exp_dir else single_group_exp_dir
    os.makedirs(os.path.join(result_dir, "a_result_folder"), exist_ok=True)
    output_path = os.path.join(result_dir, "a_result_folder", f"{result_file_name}.xlsx")

    data = {}
    chapter_groups = set()
    if multi_group_exp_dir:
        data, chapter_groups = read_group_results(multi_group_exp_dir, target_book, all_books, metrics, key_word)
        chapter_groups = sorted(chapter_groups)

    baseline_data = {}
    if single_group_exp_dir:
        for lang_pair in os.listdir(single_group_exp_dir):
            lang_pattern = re.compile(r"([\w-]+)\-([\w-]+)")
            if not lang_pattern.match(lang_pair):
                continue

            baseline_path = os.path.join(single_group_exp_dir, lang_pair)
            baseline_diff_pred = glob.glob(os.path.join(baseline_path, "diff_predictions*"))
            if baseline_diff_pred:
                baseline_data[lang_pair] = extract_diff_pred_data(baseline_diff_pred[0], metrics, target_book)
            else:
                print(f"Checking experiments under {baseline_path}...")
                sub_baseline_path = os.path.join(baseline_path, "+".join(trained_books))
                baseline_diff_pred = glob.glob(os.path.join(sub_baseline_path, "diff_predictions*"))
                if baseline_diff_pred:
                    baseline_data[lang_pair] = extract_diff_pred_data(baseline_diff_pred[0], metrics, target_book)
                else:
                    print(f"Baseline experiment has no diff_predictions file in {sub_baseline_path}")

    print("Writing data...")
    rows = flatten_dict(data, chapter_groups, metrics, baseline=baseline_data)
    create_xlsx(rows, chapter_groups, output_path, metrics)
    print(f"Result is in {output_path}")


if __name__ == "__main__":
    main()
