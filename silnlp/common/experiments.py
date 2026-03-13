import argparse
import copy
import csv
import logging
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml
from openpyxl.utils import get_column_letter

from silnlp.common.environment import SIL_NLP_ENV
from .combine_scores import is_locked
from .script_utils import is_represented, predict_script_code
from .utils import two2three_iso

EXPERIMENTS_DIR = SIL_NLP_ENV.mt_experiments_dir
SCRIPTURE_DIR = SIL_NLP_ENV.mt_scripture_dir
SAMPLE_LINES = 100
MODEL = "facebook/nllb-200"

RESULT_HEADERS = [
    "experiment",
    "steps",
    "Book",
    "BLEU",
    "chrF3",
]

METADATA_COLS = {"experiment", "language", "series", "id"}

LOGGER = logging.getLogger(__package__ + ".experiments")


def set_nested(d, dotted_key, value):
    """Set a value in a nested dict using dot notation."""
    keys = dotted_key.split(".")
    for k in keys[:-1]:
        d = d.setdefault(k, {})
    d[keys[-1]] = value


def parse_value(value):
    """Convert cell value to appropriate Python type."""
    s = str(value).strip()
    if s.lower() == "true": return True
    if s.lower() == "false": return False
    try: return int(s)
    except ValueError: pass
    try: return float(s)
    except ValueError: pass
    return s


def get_template_config(series_name, wb):
    """Read the Configs sheet and return the parsed template YAML for the given series."""
    configs = read_sheet(wb, "Configs")
    for row in configs:
        if row["series"] == series_name:
            config_path = EXPERIMENTS_DIR / row["config"]
            if not config_path.is_file():
                LOGGER.error(f"Template config not found: {config_path}")
                return None
            with open(config_path, "r", encoding="utf-8") as f:
                return yaml.safe_load(f)
    LOGGER.error(f"Series '{series_name}' not found on Configs sheet")
    return None


def build_corpus_pairs(row):
    """Build corpus_pairs list from cp1/cp2/cp3 columns in a row."""
    pairs = []
    for prefix in ["cp1", "cp2", "cp3"]:
        cp_cols = {k.replace(f"{prefix}.", ""): v for k, v in row.items()
                   if k.startswith(f"{prefix}.") and str(v).strip()}
        if not cp_cols:
            continue
        pair = {}
        for key, value in cp_cols.items():
            if key == "src":
                srcs = [s.strip() for s in str(value).split(";") if s.strip()]
                pair["src"] = srcs if len(srcs) > 1 else srcs[0]
            else:
                pair[key] = parse_value(value)
        pairs.append(pair)
    return pairs


def build_config(template, row, script_map):
    """Build a config dict from template + row overrides + auto-generated lang_codes."""
    config = copy.deepcopy(template)

    # Build corpus pairs from cp columns
    pairs = build_corpus_pairs(row)
    if pairs:
        config.setdefault("data", {})["corpus_pairs"] = pairs

    # Apply dot-notation overrides (skip metadata and cp columns)
    for key, value in row.items():
        if key in METADATA_COLS or key.startswith(("cp1.", "cp2.", "cp3.")):
            continue
        if "." not in key or not str(value).strip():
            continue
        if key == "data.lang_codes":
            LOGGER.warning("lang_codes are auto-generated — ignoring column 'data.lang_codes'")
            continue
        set_nested(config, key, parse_value(value))

    # Inject lang_codes from script_map
    lang_codes = {}
    for pair in config.get("data", {}).get("corpus_pairs", []):
        srcs = pair.get("src", [])
        if isinstance(srcs, str):
            srcs = [srcs]
        trg = pair.get("trg", "")
        for proj in srcs + ([trg] if trg else []):
            prefix = extract_prefix(proj)
            if prefix and proj in script_map:
                lang_codes[prefix] = script_map[proj]
    config.setdefault("data", {})["lang_codes"] = lang_codes

    return config


def check_scripture_files(rows):
    """Check that all scripture files referenced in cp columns exist."""
    valid, any_missing = [], []
    for row in rows:
        filenames = get_scripture_filenames(row)
        missing = [f for f in filenames if not (SCRIPTURE_DIR / f"{f}.txt").is_file()]
        if missing:
            experiment = row.get("experiment") or f"{row['language']}_{row['series']}_{row['id']}"
            any_missing.append({experiment: missing})
        else:
            valid.append(row)
    return valid, any_missing


def get_scripture_filenames(row):
    """Extract all scripture filenames from cp1/cp2/cp3 src and trg columns."""
    filenames = set()
    for prefix in ["cp1", "cp2", "cp3"]:
        src = str(row.get(f"{prefix}.src", "")).strip()
        trg = str(row.get(f"{prefix}.trg", "")).strip()
        if src:
            filenames.update(s.strip() for s in src.split(";") if s.strip())
        if trg:
            filenames.add(trg)
    return filenames


def create_series(xlsxfile, main_folder, series_name, valid_rows, script_map, relative_folder):
    """Create experiment folders and config files for a series. Updates experiment column."""
    experiment_names = []

    wb = openpyxl.load_workbook(xlsxfile)
    template = get_template_config(series_name, wb)
    if template is None:
        wb.close()
        return 1

    series_rows = [r for r in valid_rows if r["series"] == series_name]
    if not series_rows:
        LOGGER.error(f"No experiments found for series '{series_name}'")
        wb.close()
        return 1

    # Find experiment column index in the sheet
    ws = wb["Experiments"]
    headers = [str(cell.value).strip() for cell in ws[1]]
    exp_col_idx = headers.index("experiment") + 1  # 1-based for openpyxl

    for row in series_rows:
        experiment = str(row.get("experiment", "")).strip()
        if not experiment:
            experiment = f"{row['language']}_{row['series']}_{row['id']}"
        
        experiment_names.append(experiment)
        folder = main_folder / experiment
        folder.mkdir(parents=True, exist_ok=True)

        config = build_config(template, row, script_map)
        config_path = folder / "config.yml"
        with open(config_path, "w", encoding="utf-8") as f:
            yaml.dump(config, f, default_flow_style=False, sort_keys=True)
        LOGGER.info(f"Created {config_path}")

    # Update experiment column for rows that were empty
    all_rows = read_sheet(wb, "Experiments")
    for sheet_row_idx, sheet_row in enumerate(all_rows):
        if sheet_row["series"] != series_name:
            continue
        if str(sheet_row.get("experiment", "")).strip():
            continue
        experiment = f"{sheet_row['language']}_{sheet_row['series']}_{sheet_row['id']}"
        ws.cell(row=sheet_row_idx + 2, column=exp_col_idx, value=experiment)

    wb.save(xlsxfile)
    wb.close()
    LOGGER.info(f"Created {len(series_rows)} experiments for series '{series_name}'")
    
    print(f"\nHere is a command to preprocess and train the experiments:\n\n{series_command(relative_folder, experiment_names)}\n")
    return 0


def get_scripts(xlsxfile, rows, two2three_iso):
    """Return dict of filename -> lang_code. Reads cached entries from the 'scripts'
    sheet in the workbook, predicts scripts for any new filenames, and updates
    the sheet if new entries were added."""
    cache = {}

    # Read existing cache from 'scripts' sheet if it exists
    wb = openpyxl.load_workbook(xlsxfile)
    if "scripts" in wb.sheetnames:
        ws = wb["scripts"]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(header).strip() for header in next(rows_iter)]
        for row in rows_iter:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            cache[d["filename"]] = {
                "filename_iso": d["filename_iso"],
                "language_iso": d["language_iso"],
                "script": d["script"],
            }

    # Collect all filenames needed
    needed = set()
    for row in rows:
        needed.update(get_scripture_filenames(row))

    # Add missing entries
    updated = False
    for filename in sorted(needed):
        if filename in cache:
            continue
        file = SCRIPTURE_DIR / f"{filename}.txt"
        if not file.is_file():
            LOGGER.warning(f"Cannot cache script for {filename}: {file} not found")
            continue

        script_code = predict_script_code(read_scripture(file))
        if not is_represented(script_code=script_code, model=MODEL):
            if updated:
                update_sheet(xlsxfile, cache)
            LOGGER.error(f"Script {script_code} found for {file} is not known to the {MODEL} model.")

        filename_iso = extract_prefix(filename)
        language_iso = two2three_iso.get(filename_iso, filename_iso)
        if not (len(language_iso) == 3 and language_iso.isalpha()):
            LOGGER.warning(f"Skipping {filename}: language_iso '{language_iso}' is not a valid 3-letter code")
            continue

        cache[filename] = {
            "filename_iso": filename_iso,
            "language_iso": language_iso,
            "script": script_code,
        }
        LOGGER.info(f"Cached script for {filename}: {language_iso}_{script_code}")
        updated = True

    # Write back if updated
    if updated:
        update_sheet(xlsxfile, cache)
    else:
        wb.close()

    return {fn: f"{cache[fn]['language_iso']}_{cache[fn]['script']}" for fn in cache}


def read_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts, skipping empty rows."""
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows_iter)]
    rows = []
    for row in rows_iter:
        d = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        if not any(str(v).strip() for v in d.values()):
            break
        rows.append(d)
    return rows


def read_scripture(file, max_lines=SAMPLE_LINES):
    "Read non-empty, non-range lines from a scripture file"
    lines_to_skip = set(["", "<range>", "..."])
    with open(file, "r", encoding="utf-8") as file_in:
        lines = [line.strip() for line in file_in if line.strip() not in lines_to_skip]
        return "".join(lines[:max_lines])


def update_sheet(xlsxfile, cache):
    """Write out the filename, filename_isocode, 3 letter isocode and script to the 'scripts' sheet."""
    wb = openpyxl.load_workbook(xlsxfile)
    HEADERS = ["filename", "filename_iso", "language_iso", "script"]
    if "scripts" in wb.sheetnames:
        ws = wb["scripts"]
        existing = {}
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            fn = str(row[0].value).strip() if row[0].value else ""
            if fn:
                existing[fn] = row_num
        for fn in sorted(cache):
            c = cache[fn]
            vals = [fn, c["filename_iso"], c["language_iso"], c["script"]]
            if fn in existing:
                for col, val in enumerate(vals, start=1):
                    ws.cell(row=existing[fn], column=col, value=val)
            else:
                ws.append(vals)
    else:
        ws = wb.create_sheet("scripts")
        ws.append(HEADERS)
        ws.column_dimensions[get_column_letter(1)].width = 27
        ws.column_dimensions[get_column_letter(2)].width = 12
        ws.column_dimensions[get_column_letter(3)].width = 12
        ws.column_dimensions[get_column_letter(4)].width = 6
        for fn in sorted(cache):
            c = cache[fn]
            ws.append([fn, c["filename_iso"], c["language_iso"], c["script"]])
    wb.save(xlsxfile)
    wb.close()
    LOGGER.info(f"Updated scripts sheet in {xlsxfile} ({len(cache)} entries)")


def backup_workbook(wb, xlsxfile):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    backup_xlsxfile = xlsxfile.parent / f"experiments_bak_{timestamp}.xlsx"
    wb.save(backup_xlsxfile)
    wb.close()


def extract_prefix(project_name):
    """Extract the prefix (everything before the first dash) from a project name."""
    if not isinstance(project_name, str):
        return None
    match = re.match(r"^([^-]+)", project_name)
    if match:
        return match.group(1)
    return project_name


def resolve_lang_code(project_name, script_map):
    """Look up the lang_code for a project from the script cache."""
    lang_code = script_map.get(project_name)
    if not lang_code:
        raise RuntimeError(f"Could not find lang_code for {project_name} in scripts cache")
    return lang_code


def get_scripts(xlsxfile, rows, two2three_iso):
    """Return dict of filename -> lang_code. Reads cached entries from the 'scripts'
    sheet in the workbook, predicts scripts for any new filenames, and updates
    the sheet if new entries were added."""
    cache = {}
    wb = openpyxl.load_workbook(xlsxfile)
    if "scripts" in wb.sheetnames:
        ws = wb["scripts"]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(header).strip() for header in next(rows_iter)]
        for row in rows_iter:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            cache[d["filename"]] = {
                "filename_iso": d["filename_iso"],
                "language_iso": d["language_iso"],
                "script": d["script"],
            }
    needed = set()
    for row in rows:
        needed.update(get_scripture_filenames(row))
    needed.discard("")
    updated = False
    for filename in sorted(needed):
        if filename in cache:
            continue
        file = SCRIPTURE_DIR / f"{filename}.txt"
        if not file.is_file():
            LOGGER.warning(f"Cannot cache script for {filename}: {file} not found")
            continue
        script_code = predict_script_code(read_scripture(file))
        if not is_represented(script_code=script_code, model=MODEL):
            if updated:
                update_sheet(xlsxfile, cache)
            LOGGER.error(f"Script {script_code} found for {file} is not known to the {MODEL} model.")
        filename_iso = extract_prefix(filename)
        language_iso = two2three_iso.get(filename_iso, filename_iso)
        if not (len(language_iso) == 3 and language_iso.isalpha()):
            LOGGER.warning(f"Skipping {filename}: language_iso '{language_iso}' is not a valid 3-letter code")
            continue
        cache[filename] = {
            "filename_iso": filename_iso,
            "language_iso": language_iso,
            "script": script_code,
        }
        LOGGER.info(f"Cached script for {filename}: {language_iso}_{script_code}")
        updated = True
    if updated:
        update_sheet(xlsxfile, cache)
    else:
        wb.close()
    return {fn: f"{cache[fn]['language_iso']}_{cache[fn]['script']}" for fn in cache}


def check_scripture_files(rows):
    print(rows[0])
    
    """Check that all Source 1, Source 2, and Target scripture files exist."""
    valid_rows, all_missing = [], []
    for row in rows:
        language = row["language"]
        files = []
        for cp_count in range(1,4):
            files.extend([src for src in row.get(f"cp{cp_count}.src", "").split(';') if src])
            trg = row.get(f"cp{cp_count}.trg", "")
            if trg:
                files.append(trg)

            missing = [file for file in files if file and not (SCRIPTURE_DIR / f"{file}.txt").is_file()]
        if missing:
            all_missing.append({language: missing})
        else:
            valid_rows.append(row)
    return valid_rows, all_missing


def get_scores(scores_file):
    """Read scores file. Returns list of dicts with keys: Book, BLEU, chrF3.
    One dict per row (per-book + ALL). Returns empty list if file missing."""
    results = []
    with open(scores_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        bleu_col = next((c for c in reader.fieldnames if c.lower() == "bleu"), None)
        chrf_col = next((c for c in reader.fieldnames if "chrf" in c.lower()), None)
        book_col = next((c for c in reader.fieldnames if c.lower() == "book"), None)
        for row in reader:
            results.append({
                "Book": row[book_col].strip() if book_col else "",
                "BLEU": float(row[bleu_col]) if bleu_col and row[bleu_col].strip() else None,
                "chrF3": float(row[chrf_col]) if chrf_col and row[chrf_col].strip() else None,
            })
    return results


def collect_results(xlsxfile, main_folder, valid_rows, overwrite):
    """Collect results from all experiments and write to 'results' sheet."""
    wb = openpyxl.load_workbook(xlsxfile)
    if overwrite:
        backup_workbook(wb, xlsxfile)
        wb = openpyxl.load_workbook(xlsxfile)

    existing = {}
    if not overwrite and "results" in wb.sheetnames:
        for r in read_sheet(wb, "results"):
            key = (r["experiment"], r.get("steps", ""), r.get("Book", ""))
            existing[key] = r

    all_results = []
    for row in valid_rows:
        experiment = str(row.get("experiment", "")).strip()
        if not experiment:
            experiment = f"{row['language']}_{row['series']}_{row['id']}"

        folder = main_folder / experiment
        if not folder.is_dir():
            LOGGER.warning(f"Folder not found: {experiment}")
            continue

        scores_files = sorted(folder.glob("scores-*.csv"))
        if not scores_files:
            LOGGER.warning(f"No scores files found in {experiment}")
            continue

        for scores_file in scores_files:
            match = re.search(r"scores-(\d+)\.csv", scores_file.name)
            steps = match.group(1) if match else ""

            sample_key = (experiment, steps, "ALL")
            ex = existing.get(sample_key, {})
            need_scores = overwrite or not ex.get("BLEU")

            if need_scores:
                scores = get_scores(scores_file)
            else:
                scores = [
                    {"Book": r["Book"], "BLEU": r["BLEU"], "chrF3": r["chrF3"]}
                    for k, r in existing.items()
                    if k[0] == experiment and k[1] == steps
                ]

            for s in scores:
                all_results.append({
                    "experiment": experiment,
                    "steps": steps,
                    "Book": s["Book"],
                    "BLEU": s["BLEU"],
                    "chrF3": s["chrF3"],
                })

    if "results" in wb.sheetnames:
        del wb["results"]
    ws = wb.create_sheet("results")
    ws.append(RESULT_HEADERS)
    for r in all_results:
        ws.append([r.get(h) for h in RESULT_HEADERS])

    wb.save(xlsxfile)
    wb.close()
    return all_results


def create_analysis_sheet(xlsxfile):
    """Read 'results' sheet and create a single 'analysis' sheet with deltas for ALL books only."""
    wb = openpyxl.load_workbook(xlsxfile)
    results = read_sheet(wb, "results")
    # TODO: Phase 2+ will define how experiments are grouped for comparison.
    # For now this is a placeholder.
    wb.save(xlsxfile)
    wb.close()


def series_command(relative_folder, experiments):
    """Create the command that will preprocess and run the experiments."""
    loop = f"for exp in {' '.join(experiments)}; do" 
    prep = f"poetry run python -m silnlp.nmt.preprocess --stats {relative_folder}" + "/${exp}"
    train= f"poetry run python -m silnlp.nmt.experiment --save-checkpoints --save-confidences --clearml-queue jobs_backlog --clearml-tag eitl --train --test --score-by-book {relative_folder}" + "/${exp}"
    return loop + f"\n  echo {prep}\n  {prep}\n  echo {train}\n  {train}\ndone"


def discover_experiments(xlsxfile, experiments_dir):
    """Discover experiment folders with effective-config, infer folder, and scores files.
    Write results to a new xlsx workbook."""
    SKIP_FOLDERS = {"align", "analyze", "analyse", "alignment", "alignments"}
    REQUIRED_MODEL = "facebook/nllb-200-distilled-1.3B"
    REQUIRED_SEED = 111

    configs = sorted(experiments_dir.rglob("effective-config-*.yml"))
    print(f"Found {len(configs)} effective-config files. Filtering...")

    rows = []
    skipped = {"no_infer": 0, "no_scores": 0, "wrong_model": 0, "wrong_seed": 0,
               "too_many_cp": 0, "bad_folder": 0, "parse_error": 0}

    for i, config_path in enumerate(configs):
        if (i + 1) % 500 == 0:
            print(f"  Processed {i + 1}/{len(configs)} ({len(rows)} kept so far)")

        folder = config_path.parent

        # Skip excluded folder names
        if any(part.lower() in SKIP_FOLDERS for part in folder.parts):
            skipped["bad_folder"] += 1
            continue

        # Must have infer folder
        if not (folder / "infer").is_dir():
            skipped["no_infer"] += 1
            continue

        # Must have scores files
        scores_files = sorted(folder.glob("scores-*.csv"))
        if not scores_files:
            skipped["no_scores"] += 1
            continue

        # Parse config
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = yaml.safe_load(f)
        except Exception as e:
            LOGGER.warning(f"Failed to parse {config_path}: {e}")
            skipped["parse_error"] += 1
            continue

        # Filter model and seed
        if config.get("model") != REQUIRED_MODEL:
            skipped["wrong_model"] += 1
            continue
        data = config.get("data", {})
        if data.get("seed") != REQUIRED_SEED:
            skipped["wrong_seed"] += 1
            continue

        # Check corpus_pairs
        corpus_pairs = data.get("corpus_pairs", [])
        if len(corpus_pairs) > 3:
            skipped["too_many_cp"] += 1
            continue

        # Build row
        experiment = str(folder.relative_to(experiments_dir))
        lang_codes = data.get("lang_codes", {})
        lang_codes_str = ";".join(f"{k}:{v}" for k, v in sorted(lang_codes.items()))

        row = {"experiment": experiment, "lang_codes": lang_codes_str}

        for idx, cp in enumerate(corpus_pairs, start=1):
            prefix = f"cp{idx}"
            src = cp.get("src", "")
            if isinstance(src, list):
                src = ";".join(src)
            row[f"{prefix}.type"] = cp.get("type", "")
            row[f"{prefix}.mapping"] = cp.get("mapping", "")
            row[f"{prefix}.src"] = src
            row[f"{prefix}.trg"] = cp.get("trg", "")
            row[f"{prefix}.corpus_books"] = cp.get("corpus_books", "")
            row[f"{prefix}.test_books"] = cp.get("test_books", "")

        rows.append(row)

    # Collect all column headers from all rows
    all_keys = []
    seen = set()
    for key in ["experiment", "lang_codes"]:
        all_keys.append(key)
        seen.add(key)
    for row in rows:
        for k in row:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # Write xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discovered_Experiments"
    ws.append(all_keys)
    for row in rows:
        ws.append([row.get(h, "") for h in all_keys])

    wb.save(xlsxfile)
    wb.close()

    print(f"\nDone. Kept {len(rows)} experiments, wrote to {xlsxfile}")
    print(f"Skipped: {skipped}")


def main():
    parser = argparse.ArgumentParser(description="Manage NMT experiments.")
    parser.add_argument("folder", help="Root experiment folder name (relative to mt_experiments_dir).")
    parser.add_argument("--xlsxfile", default="experiments.xlsx", help="Workbook filename.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing results.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--collect-scripts", action="store_true", help="Update the scripts sheet.")
    group.add_argument("--create-series", type=str, metavar="SERIES", help="Create experiment configs for the named series.")
    group.add_argument("--collect-results", action="store_true", help="Collect the results of the experiments.")
    group.add_argument("--discover", type=str, help="Discover existing config files and store the results.")

    args = parser.parse_args()

    main_folder = EXPERIMENTS_DIR / args.folder
    xlsxfile = main_folder / args.xlsxfile
    xlsxfile_rel = xlsxfile.relative_to(EXPERIMENTS_DIR)
    relative_folder = xlsxfile_rel.parent

    if not xlsxfile.is_file():
        LOGGER.error(f"\nExperiment workbook not found: {xlsxfile}.")
        return 1

    if lockfile := is_locked(xlsxfile):
        print(f"Found lock file: {lockfile}")
        print(f"Please close {xlsxfile.name} or delete the lock file and try again.")
        return 1

    if args.discover:
        # Folder to search 
        search_folder = EXPERIMENTS_DIR / args.discover
        discover_xlsxfile = main_folder / "discovered_experiments.xlsx"
        if not search_folder.is_dir():
            LOGGER.error(f"\nCouldn't find the folder to search: {search_folder}.")
            return 1
        else:
            discover_experiments(discover_xlsxfile, search_folder)
            return 0

    wb = openpyxl.load_workbook(xlsxfile)
    rows = read_sheet(wb, "Experiments")
    wb.close()

    LOGGER.info(f"Read {len(rows)} experiment definitions from {xlsxfile}")
    valid_rows, any_missing = check_scripture_files(rows)

    if any_missing:
        LOGGER.warning("Missing scripture files:")
        for entry in any_missing:
            for exp, missing in entry.items():
                print(f"  {exp}: {missing}")
        return 1

    LOGGER.info(f"All scripture files present for {len(valid_rows)} experiments.")

    script_map = get_scripts(xlsxfile, valid_rows, two2three_iso)
    if not script_map:
        LOGGER.error("Could not determine scripts for any projects.")
        return 1

    if args.collect_scripts:
        return 0

    if args.create_series:
        return create_series(xlsxfile, main_folder, args.create_series, valid_rows, script_map, relative_folder)

    if args.collect_results:
        results = collect_results(xlsxfile, main_folder, valid_rows, args.overwrite)
        LOGGER.info(f"Wrote {len(results)} result rows to {xlsxfile}")

    return 0


if __name__ == "__main__":
    main()

