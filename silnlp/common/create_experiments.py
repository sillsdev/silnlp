import argparse
import csv
import logging
import re
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import yaml
from openpyxl.chart import BarChart, Reference
from scipy.stats import spearmanr, wilcoxon

from silnlp.common.environment import SIL_NLP_ENV

from .combine_scores import is_locked
from .script_utils import is_represented, predict_script_code
from .utils import two2three_iso

EXPERIMENTS_DIR = SIL_NLP_ENV.mt_experiments_dir
SCRIPTURE_DIR = SIL_NLP_ENV.mt_scripture_dir
SAMPLE_LINES = 100
MODEL = "facebook/nllb-200"
RESULT_HEADERS = [
    "Target_language",
    "Mapping",
    "Source 1",
    "Source 2",
    "Target",
    "train_lines",
    "unique_train_lines",
    "test_lines",
    "src_mean_chars_per_token",
    "trg_mean_chars_per_token",
    "src_mean_tokens_per_verse",
    "trg_mean_tokens_per_verse",
    "Book",
    "BLEU",
    "chrF3++",
]

LOGGER = logging.getLogger(__package__ + ".create_experiments")


def read_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts, skipping empty rows."""
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows_iter)]
    rows = []
    for row in rows_iter:
        d = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        if not any(str(v).strip() for v in d.values()):
            break
        rows.append(d)
    return rows


def read_scripture(file, max_lines=SAMPLE_LINES):
    "Read non-empty, non-range lines from a scripture file"
    lines_to_skip = set(["", "<range>", "..."])
    with open(file, "r", encoding="utf-8") as file_in:
        lines = [line.strip() for line in file_in if line.strip() not in lines_to_skip]
        return "".join(lines[:max_lines])


def update_sheet(wb, workbook_path, cache):
    """Write out the filename, filename_isocode, 3 letter isocode and script to the "scripts" sheet in the spreadsheet"""
    HEADERS = ["filename", "filename_iso", "language_iso", "script"]
    if "scripts" in wb.sheetnames:
        ws = wb["scripts"]
        # Build a map of filename -> row number for existing entries
        existing = {}
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            fn = str(row[0].value).strip() if row[0].value else ""
            if fn:
                existing[fn] = row_num

        for fn in sorted(cache):
            c = cache[fn]
            vals = [fn, c["filename_iso"], c["language_iso"], c["script"]]
            if fn in existing:
                for col, val in enumerate(vals, start=1):
                    ws.cell(row=existing[fn], column=col, value=val)
            else:
                ws.append(vals)
    else:
        ws = wb.create_sheet("scripts")
        ws.append(HEADERS)
        for fn in sorted(cache):
            c = cache[fn]
            ws.append([fn, c["filename_iso"], c["language_iso"], c["script"]])

    wb.save(workbook_path)
    LOGGER.info(f"Updated scripts sheet in {workbook_path} ({len(cache)} entries)")


def backup_workbook(wb, workbook_file):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    backup_workbook_file = workbook_file.parent / f"experiments_bak_{timestamp}.xlsx"
    wb.save(backup_workbook_file)


def get_scripts(workbook_path, rows, two2three_iso):
    """Return dict of filename -> lang_code. Reads cached entries from the 'scripts'
    sheet in the workbook, predicts scripts for any new filenames, and updates
    the sheet if new entries were added."""
    cache = {}

    # Read existing cache from 'scripts' sheet if it exists
    wb = openpyxl.load_workbook(workbook_path)
    if "scripts" in wb.sheetnames:
        ws = wb["scripts"]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(header).strip() for header in next(rows_iter)]
        for row in rows_iter:
            d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            cache[d["filename"]] = {
                "filename_iso": d["filename_iso"],
                "language_iso": d["language_iso"],
                "script": d["script"],
            }

    # Collect all filenames needed
    needed = set()
    for row in rows:
        needed.update([row["Source 1"], row["Source 2"], row["Target"]])
    needed.discard("")

    # Add missing entries
    updated = False
    for filename in sorted(needed):
        if filename in cache:
            continue
        file = SCRIPTURE_DIR / f"{filename}.txt"
        if not file.is_file():
            LOGGER.warning(f"Cannot cache script for {filename}: {file} not found")
            continue

        script_code = predict_script_code(read_scripture(file))
        if not is_represented(script_code=script_code, model=MODEL):
            if updated:
                update_sheet(wb, workbook_path, cache)
            LOGGER.error(f"Script {script_code} found for {file} is not known to the {MODEL} model.")

        filename_iso = extract_prefix(filename)
        language_iso = two2three_iso.get(filename_iso, filename_iso)
        if not (len(language_iso) == 3 and language_iso.isalpha()):
            LOGGER.warning(f"Skipping {filename}: language_iso '{language_iso}' is not a valid 3-letter code")
            continue

        cache[filename] = {
            "filename_iso": filename_iso,
            "language_iso": language_iso,
            "script": script_code,
        }
        LOGGER.info(f"Cached script for {filename}: {language_iso}_{script_code}")
        updated = True

    # Write back if updated
    if updated:
        update_sheet(wb, workbook_path, cache)
    else:
        wb.close()

    return {fn: f"{cache[fn]['language_iso']}_{cache[fn]['script']}" for fn in cache}


def extract_prefix(project_name):
    """Extract the prefix (everything before the first dash) from a project name."""
    if not isinstance(project_name, str):
        return None
    match = re.match(r"^([^-]+)", project_name)
    if match:
        return match.group(1)
    return project_name


def resolve_lang_code(project_name, script_map):
    """Look up the lang_code for a project from the script cache."""
    lang_code = script_map.get(project_name)
    if not lang_code:
        raise RuntimeError(f"Could not find lang_code for {project_name} in scripts cache")
    return lang_code


def create_alignment_config(folder, rows):
    print(f"Creating Alignment config. Run the alignments before continuing.")
    all_src = set()
    all_trg = set()
    for row in rows:
        all_src.add(row["Source 1"])
        if row["Source 2"]:
            all_src.add(row["Source 2"])
        all_trg.add(row["Target"])

    config = {
        "data": {
            "aligner": "eflomal",
            "corpus_pairs": [
                {
                    "type": "train",
                    "src": sorted(list(all_src)),
                    "trg": sorted(list(all_trg)),
                    "mapping": "many_to_many",
                    "test_size": 0,
                    "val_size": 0,
                }
            ],
        }
    }

    align_dir = folder / "Align"
    align_dir.mkdir(exist_ok=True)
    with open(align_dir / "config.yml", "w", encoding="utf-8") as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False)
    LOGGER.info(f"Created alignment config: {align_dir / 'config.yml'}")
    exit(0)


def check_scripture_files(rows):
    """Check that all Source 1, Source 2, and Target scripture files exist. Returns list of valid rows."""
    valid, any_missing = [], []
    for row in rows:
        lang, src1, src2, trg = row["Target_language"], row["Source 1"], row["Source 2"], row["Target"]
        missing = [file for file in (src1, src2, trg) if not (SCRIPTURE_DIR / f"{file}.txt").is_file()]
        if missing:
            any_missing.append({lang: missing})
        else:
            valid.append(row)
    return valid, any_missing


def create_config(mapping_type, lang_codes, src_list, trg, corpus_books, test_books):
    config = {
        "data": {
            "corpus_pairs": [
                {
                    "type": "train",
                    "corpus_books": corpus_books,
                    "mapping": mapping_type,
                    "src": src_list,
                    "trg": trg,
                },
                {
                    "type": "test",
                    "corpus_books": test_books,
                    "src": src_list[0],
                    "trg": trg,
                },
            ],
            "lang_codes": lang_codes,
            "seed": 111,
            "tokenizer": {"update_src": True, "update_trg": True},
        },
        # "eval": {"early_stopping": None, "eval_steps": 1000, 'eval_strategy': 'no'},
        "model": "facebook/nllb-200-distilled-1.3B",
        # "train": {"max_steps": 7000, "save_steps": 5000, "save_strategy": "steps", "save_total_limit": 1},
    }
    return config


def write_config_file(row, main_folder, overwrite):

    language = row["Target_language"]
    series = row["Series"]
    src1 = row["Source 1"]
    src2 = row["Source 2"]
    trg = row["Target"]
    corpus_books = row["corpus_books"]
    test_books = row["test_books"]

    experiments = [
        #("single", "one_to_one", [src1]),
        ("mixed", "mixed_src", [src1, src2]),
        ("many", "many_to_many", [src1, src2]),
    ]

    for suffix, mapping_type, src_list in experiments:
        # if suffix != "single" and not src2:
        #     continue
        experiment_name = f"{language}_{series}_{suffix}"
        experiment_folder = main_folder / experiment_name
        experiment_folder.mkdir(exist_ok=True)

        config_file = experiment_folder / "config.yml"
        if config_file.is_file() and not overwrite:
            LOGGER.info(f"Skipping existing config: {config_file}")
            continue

        # lang_codes
        lang_codes = {}
        # We need to resolve all prefixes: src1, trg, and src2 (if it exists)
        projects_to_resolve = [src1, trg]
        if src2:
            projects_to_resolve.append(src2)

        for proj in projects_to_resolve:
            prefix = extract_prefix(proj)
            lang_codes[prefix] = resolve_lang_code(proj, script_map)

            if not lang_codes[prefix]:
                raise RuntimeError(
                    f"Could not find lang_code for {prefix} for {project_name}. Not present on scripts sheet in {workbook_file}."
                )

        config = create_config(mapping_type, lang_codes, src_list, trg, corpus_books, test_books)

        with open(config_file, "w", encoding="utf-8") as cf:
            yaml.dump(config, cf, default_flow_style=False, sort_keys=False)

        LOGGER.info(f"Created experiment config: {config_file}")


def count_lines_and_unique_lines(file):
    """Count the total and number of unique lines in a file."""
    if not file.is_file():
        LOGGER.warning(f"Couldn't find train.vref file {file}")
        return None, None

    with open(file, mode="r", encoding="utf-8") as f:
        lines = f.readlines()
    return len(lines), len(set(lines))


def get_tokenization_stats(stats_file):
    """Read Mean Tokens/Verse and Mean Characters/Token for Source and Target
    from tokenization_stats.csv. Returns dict with four keys, or empty dict if file missing."""

    with open(stats_file, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # skip category header row
        next(reader)  # skip column names row
        result = {}
        for row in reader:
            if not row or not row[0].strip():
                continue
            side = row[0].strip()
            if side == "Source":
                result["src_mean_tokens_per_verse"] = float(row[5])
                result["src_mean_chars_per_token"] = float(row[17])
            elif side == "Target":
                result["trg_mean_tokens_per_verse"] = float(row[5])
                result["trg_mean_chars_per_token"] = float(row[17])
    return result


def get_scores(scores_file):
    """Read scores file. Returns list of dicts with keys: Book, BLEU, chrF3++.
    One dict per row (per-book + ALL). Returns empty list if file missing."""
    results = []
    with open(scores_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        bleu_col = next((c for c in reader.fieldnames if c.lower() == "bleu"), None)
        chrf_col = next((c for c in reader.fieldnames if "chrf" in c.lower()), None)
        book_col = next((c for c in reader.fieldnames if c.lower() == "book"), None)
        for row in reader:
            results.append(
                {
                    "Book": row[book_col].strip() if book_col else "",
                    "BLEU": float(row[bleu_col]) if bleu_col and row[bleu_col].strip() else None,
                    "chrF3++": float(row[chrf_col]) if chrf_col and row[chrf_col].strip() else None,
                }
            )
    return results


def collect_results(wb, main_folder, valid_rows, workbook_file, overwrite):
    """Collect results from all experiments and write to 'results' sheet.
    If not overwrite, reads existing results and only fetches data for missing groups."""
    if overwrite:
        backup_workbook_file(wb, workbook_file)

    count_cached, count_read = 0, 0

    # Build index of existing results keyed by (Target_language, Mapping, Book)
    existing = {}
    if not overwrite and "results" in wb.sheetnames:
        for r in read_sheet(wb, "results"):
            key = (r["Target_language"], r["Series"], r["Mapping"], r.get("Book", ""))
            existing[key] = r

    all_results = []
    for row in valid_rows:
        language = row["Target_language"]
        series = row["Series"]
        src1, src2, trg = row["Source 1"], row["Source 2"], row["Target"]

        for suffix, mapping in [("many", "many_to_many"), ("mixed", "mixed_src"), ("single", "one_to_one")]:
            folder = main_folder / f"{language}_{series}_{suffix}"
            if not folder.is_dir():
                LOGGER.warning(f"Folder not found: {folder}")
                continue

            train_vref_file, test_vref_file, stats_file, scores_file = (
                folder / file
                for file in ["train.vref.txt", "test.vref.txt", "tokenization_stats.csv", "scores-5000.csv"]
            )
            preprocess_files = [stats_file, train_vref_file, test_vref_file]
            missing_preprocess_files = [f for f in preprocess_files if not f.is_file()]
            if missing_preprocess_files:
                LOGGER.info(f"These preprocess files are missing, skipping this experiment. {missing_preprocess_files}")
                continue

            # Check what we already have for this experiment
            sample_key = (language, series, mapping, "ALL")
            ex = existing.get(sample_key, {})

            # Group 1: config — always available from the row
            # Group 2: scores
            need_scores = overwrite or not ex.get("BLEU")
            # Group 3: vref line counts
            need_lines = overwrite or not ex.get("train_lines")
            # Group 4: tokenization stats
            need_tok = overwrite or not ex.get("src_mean_chars_per_token")

            if need_lines:
                train_lines, unique_train_lines = count_lines_and_unique_lines(train_vref_file)
                test_lines, _ = count_lines_and_unique_lines(test_vref_file)
            else:
                train_lines = ex["train_lines"]
                unique_train_lines = ex["unique_train_lines"]
                test_lines = ex["test_lines"]

            if need_tok:
                tok_stats = get_tokenization_stats(stats_file)
            else:
                tok_stats = {
                    "src_mean_chars_per_token": ex["src_mean_chars_per_token"],
                    "trg_mean_chars_per_token": ex["trg_mean_chars_per_token"],
                    "src_mean_tokens_per_verse": ex["src_mean_tokens_per_verse"],
                    "trg_mean_tokens_per_verse": ex["trg_mean_tokens_per_verse"],
                }

            if need_scores:
                if scores_file.is_file():
                    scores = get_scores(scores_file)
                else:
                    scores = [{"Book": "", "BLEU": None, "chrF3++": None}]
                    LOGGER.warning(f"{scores_file} not found in {folder}")
            else:
                # Reconstruct scores from existing rows
                scores = [
                    {"Book": r["Book"], "BLEU": r["BLEU"], "chrF3++": r["chrF3++"]}
                    for k, r in existing.items()
                    if k[0] == language and k[1] == mapping
                ]

            for s in scores:
                all_results.append(
                    [
                        language,
                        series,
                        mapping,
                        src1,
                        src2,
                        trg,
                        train_lines,
                        unique_train_lines,
                        test_lines,
                        tok_stats.get("src_mean_chars_per_token"),
                        tok_stats.get("trg_mean_chars_per_token"),
                        tok_stats.get("src_mean_tokens_per_verse"),
                        tok_stats.get("trg_mean_tokens_per_verse"),
                        s["Book"],
                        s["BLEU"],
                        s["chrF3++"],
                    ]
                )
            if not need_scores and not need_lines and not need_tok:
                count_cached += 1
            else:
                count_read += 1
            # print(
            #    f"Found results for {language}_{mapping} (scores:{'read' if need_scores else 'cached'} lines:{'read' if need_lines else 'cached'} tok:{'read' if need_tok else 'cached'})"
            # )

    print(
        f"Found cached results for {count_cached} experiments and attempted to collect results for {count_read} experiments."
    )

    # Write to results sheet
    if "results" in wb.sheetnames:
        del wb["results"]
    ws = wb.create_sheet("results")
    ws.append(RESULT_HEADERS)
    for r in all_results:
        ws.append(r)

    return wb, all_results


def create_analysis_sheets(wb):
    """Read the 'results' sheet and create per-book analysis sheets with deltas."""
    results = read_sheet(wb, "results")

    # Group by (Target_language, Book, Mapping)
    data = {}
    for r in results:
        key = (r["Target_language"], r["Series"], r["Book"])
        if key not in data:
            data[key] = {}
        data[key][r["Mapping"]] = r

    books = set(r["Book"] for r in results)
    books.discard(None)
    books.discard("")

    # Find all books
    books = sorted(books)

    ANALYSIS_HEADERS = [
        "Target_language",
        "Series",
        "BLEU_one_to_one",
        "BLEU_mixed_src",
        "BLEU_many_to_many",
        "chrF3++_one_to_one",
        "chrF3++_mixed_src",
        "chrF3++_many_to_many",
        "train_lines",
        "ratio_chars_per_token",
        "ratio_tokens_per_verse",
        "diff_tokens_per_verse",
        "BLEU_delta_m2m_mix",
        "BLEU_delta_mix_o2o",
        "chrF3++_delta_m2m_mix",
        "chrF3++_delta_mix_o2o",
    ]

    for book in books:
        sheet_name = f"analysis_{book}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_out = wb.create_sheet(sheet_name)
        ws_out.append(ANALYSIS_HEADERS)

        # Get all languages that have data for this series and book.
        langs = sorted(set(lang for lang, series, bk in data if series series and bk == book ))

        for lang in langs:
            mappings = data.get((lang, series book), {})
            m2m = mappings.get("many_to_many", {})
            mix = mappings.get("mixed_src", {})
            o2o = mappings.get("one_to_one", {})

            bleu_m2m = m2m.get("BLEU")
            bleu_mix = mix.get("BLEU")
            bleu_o2o = o2o.get("BLEU")
            chrf_m2m = m2m.get("chrF3++")
            chrf_mix = mix.get("chrF3++")
            chrf_o2o = o2o.get("chrF3++")

            # Use many_to_many stats, fall back to mixed_src
            stats = m2m or mix
            src_cpt = stats.get("src_mean_chars_per_token")
            trg_cpt = stats.get("trg_mean_chars_per_token")
            src_tpv = stats.get("src_mean_tokens_per_verse")
            trg_tpv = stats.get("trg_mean_tokens_per_verse")

            ratio_cpt = src_cpt / trg_cpt if src_cpt and trg_cpt else None
            ratio_tpv = src_tpv / trg_tpv if src_tpv and trg_tpv else None
            diff_tpv = src_tpv - trg_tpv if src_tpv and trg_tpv else None

            def delta(a, b):
                return round(a - b, 4) if a is not None and b is not None else None

            ws_out.append(
                [
                    lang,
                    bleu_o2o,
                    bleu_mix,
                    bleu_m2m,
                    chrf_o2o,
                    chrf_mix,
                    chrf_m2m,
                    stats.get("train_lines"),
                    round(ratio_cpt, 4) if ratio_cpt else None,
                    round(ratio_tpv, 4) if ratio_tpv else None,
                    round(diff_tpv, 4) if diff_tpv else None,
                    delta(bleu_m2m, bleu_mix),
                    delta(bleu_mix, bleu_o2o),
                    delta(chrf_m2m, chrf_mix),
                    delta(chrf_mix, chrf_o2o),
                ]
            )

    return wb


def create_summary_sheet(wb):
    """Create a 'summary' sheet with mean, median, +ve/−ve counts and Wilcoxon p-values
    for each book's BLEU and chrF3++ deltas, read from the analysis sheets."""

    SUMMARY_HEADERS = [
        "Book",
        "Metric",
        "Delta",
        "Mean",
        "Median",
        "Count +ve",
        "Count −ve",
        "n",
        "p-value",
    ]

    # Find all analysis sheets
    analysis_sheets = sorted([s for s in wb.sheetnames if s.startswith("analysis_")])

    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws = wb.create_sheet("summary")
    ws.append(SUMMARY_HEADERS)

    # Column indices (0-based) in analysis sheets for the four delta columns
    delta_cols = {
        ("BLEU", "m2m_mix"): 11,
        ("BLEU", "mix_o2o"): 12,
        ("chrF3++", "m2m_mix"): 13,
        ("chrF3++", "mix_o2o"): 14,
    }

    for sheet_name in analysis_sheets:
        book = sheet_name.replace("analysis_", "")
        ws_in = wb[sheet_name]

        # Read all data rows (skip header)
        all_rows = list(ws_in.iter_rows(min_row=2, values_only=True))

        for (metric, delta_type), col_idx in delta_cols.items():
            values = [
                row[col_idx] for row in all_rows if row[col_idx] is not None and isinstance(row[col_idx], (int, float))
            ]

            n = len(values)
            if n == 0:
                ws.append([book, metric, delta_type, None, None, None, None, 0, None])
                continue

            mean = round(sum(values) / n, 4)
            median = (
                round(sorted(values)[n // 2], 4)
                if n % 2 == 1
                else round((sorted(values)[n // 2 - 1] + sorted(values)[n // 2]) / 2, 4)
            )
            count_pos = sum(1 for v in values if v > 0)
            count_neg = sum(1 for v in values if v < 0)

            # Wilcoxon requires at least 10 non-zero differences for a meaningful test
            non_zero = [v for v in values if v != 0]
            if len(non_zero) >= 10:
                _, p_value = wilcoxon(non_zero)
                p_value = round(p_value, 6)
            else:
                p_value = None

            ws.append([book, metric, delta_type, mean, median, count_pos, count_neg, n, p_value])
    return wb


def create_correlation_sheet(wb):
    "Compute Spearman correlations between score deltas and experiment variables per book."

    analysis_sheets = sorted([s for s in wb.sheetnames if s.startswith("analysis_")])

    headers = ["Book", "Metric", "Variable", "rho", "p-value", "n"]
    if "correlations" in wb.sheetnames:
        del wb["correlations"]
    ws = wb.create_sheet("correlations")
    ws.append(headers)

    var_cols = ["train_lines", "ratio_chars_per_token", "ratio_tokens_per_verse", "diff_tokens_per_verse"]
    delta_cols = dict(BLEU="BLEU_delta_m2m_mix", chrF3pp="chrF3++_delta_m2m_mix")

    for sheet_name in analysis_sheets:
        book = sheet_name.replace("analysis_", "")
        rows = read_sheet(wb, sheet_name)
        for metric, dcol in delta_cols.items():
            for var in var_cols:
                pairs = [
                    (float(r[dcol]), float(r[var]))
                    for r in rows
                    if r.get(dcol) not in (None, "") and r.get(var) not in (None, "")
                ]
                if len(pairs) < 5:
                    ws.append([book, metric, var, None, None, len(pairs)])
                    continue
                deltas, vals = zip(*pairs)
                rho, p = spearmanr(deltas, vals)
                ws.append([book, metric, var, round(rho, 4), round(p, 6), len(pairs)])
    return wb


def add_charts_to_workbook(wb):
    rows = read_sheet(wb, "summary")
    m2m = [r for r in rows if r["Delta"] == "m2m_mix"]

    for metric in ["BLEU", "chrF3++"]:
        data = [r for r in m2m if r["Metric"] == metric]
        sheet_name = f"chart_{metric.replace('+', 'p')}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(["Book", "Mean Delta"])
        for r in data:
            ws.append([r["Book"], float(r["Mean"])])

        chart = BarChart()
        chart.type = "bar"
        chart.title = f"{metric}: many_to_many − mixed_src"
        chart.x_axis.title = "Mean Delta"
        chart.y_axis.title = "Book"
        chart.x_axis.crosses = "autoZero"
        chart.style = 10
        chart.width = 20
        chart.height = max(8, len(data) * 1.5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
        vals = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "D2")

    return wb


def plot_diverging_deltas(wb, output_folder=None):
    rows = read_sheet(wb, "summary")

    m2m = [r for r in rows if r["Delta"] == "m2m_mix"]

    for metric in ["BLEU", "chrF3++"]:
        data = [r for r in m2m if r["Metric"] == metric]
        books = [r["Book"] for r in data]
        means = [float(r["Mean"]) for r in data]
        pvals = [r["p-value"] for r in data]
        sig = [p is not None and p != "" and float(p) < 0.05 for p in pvals]
        colors = ["#2ecc71" if s else "#95a5a6" for s in sig]

        y = np.arange(len(books))
        fig, ax = plt.subplots(figsize=(10, max(4, len(books) * 0.5)))
        ax.barh(y, means, color=colors, edgecolor="white", height=0.6)
        ax.axvline(0, color="black", linewidth=0.8)
        ax.set_yticks(y)
        ax.set_yticklabels(books)
        ax.set_xlabel(f"{metric} delta (many_to_many − mixed_src)")
        ax.set_title(f"{metric}: many_to_many vs mixed_src by book\n(green = p < 0.05)")
        ax.invert_yaxis()
        plt.tight_layout()
        if output_folder:
            plt.savefig(output_folder / f"delta_{metric.replace('+', 'p')}.png", dpi=150)
        plt.show()


def create_correlation_sheet(wb):
    "Compute Spearman correlations between score deltas and experiment variables per book."
    analysis_sheets = sorted([s for s in wb.sheetnames if s.startswith("analysis_")])

    headers = ["Book", "Metric", "Variable", "rho", "p-value", "n"]
    if "correlations" in wb.sheetnames:
        del wb["correlations"]
    ws = wb.create_sheet("correlations")
    ws.append(headers)

    var_cols = ["train_lines", "ratio_chars_per_token", "ratio_tokens_per_verse", "diff_tokens_per_verse"]
    delta_cols = dict(BLEU="BLEU_delta_m2m_mix", chrF3pp="chrF3++_delta_m2m_mix")

    for sheet_name in analysis_sheets:
        book = sheet_name.replace("analysis_", "")
        rows = read_sheet(wb, sheet_name)
        for metric, dcol in delta_cols.items():
            for var in var_cols:
                pairs = [
                    (float(r[dcol]), float(r[var]))
                    for r in rows
                    if r.get(dcol) not in (None, "") and r.get(var) not in (None, "")
                ]
                if len(pairs) < 5:
                    ws.append([book, metric, var, None, None, len(pairs)])
                    continue
                deltas, vals = zip(*pairs)
                rho, p = spearmanr(deltas, vals)
                ws.append([book, metric, var, round(rho, 4), round(p, 6), len(pairs)])
    return wb


def main():
    parser = argparse.ArgumentParser(description="Create NLLB experiment configurations with alignment and templates.")
    parser.add_argument("folder", help="Root experiment folder name (relative to mt_experiments_dir).")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing experiment configs or results.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--create", type=str, help="Create a series of experiment folders with their config.yml files.")
    group.add_argument("--collect-scripts", action="store_true", help="Update the scripts sheet.")
    group.add_argument("--collect-results", action="store_true", help="Collect the results of the experiments.")
    group.add_argument("--analyze", action="store_true", help="Analyse the results.")
    group.add_argument("--collect-and-analyze", action="store_true", help="Collect results and run analysis.")

    args = parser.parse_args()

    main_folder = EXPERIMENTS_DIR / args.folder
    workbook_file = main_folder / "experiments.xlsx"

    if not workbook_file.is_file():
        LOGGER.error(
            f"\nExperiment workbook not found: {workbook_file}. This spreadsheet is required to define the experiments to be run."
        )
        return 1

    # Check for lock files and ask the user to close them.
    if lockfile := is_locked(workbook_file):
        print(f"Found lock file: {lockfile}")
        print(
            f"Please close {workbook_file.name} in folder {workbook_file.parent} OR if it is closed, delete the lock file and try again."
        )
        return 1

    wb = openpyxl.load_workbook(workbook_file)
    rows = read_sheet(wb, "experiments")

    if args.create:
        rows = [row for row in rows if row["Series"] == args.create]
    
    LOGGER.info(f"Read {len(rows)} experiment definitions from the 'experiments' sheet in {workbook_file}")
    valid_rows, any_missing = check_scripture_files(rows)

    if valid_rows and not any_missing:
        LOGGER.info(f"All the scripture files required for the {len(valid_rows)} experiments exist.")
    else:
        LOGGER.warning(f"The following files are missing for these experiments:")
        for lang, missing in any_missing.items():
            print(f"{lang}   : {missing}")

        return 1
    if args.create or args.collect_scripts:
        script_map = get_scripts(workbook_file, valid_rows, two2three_iso)
        if not script_map:
            LOGGER.error(f"\nCould not determine scripts for any projects.")
            exit(1)
        if args.collect_scripts:
            exit(0)

    if args.create:
        for row in valid_rows:
            write_config_file(row, main_folder, args.overwrite)

    if args.collect_results or args.collect_and_analyze:
        wb, results = collect_results(wb, main_folder, valid_rows, workbook_file, args.overwrite)
        wb.save(workbook_file)
        LOGGER.info(f"Wrote {len(results)} result rows to {workbook_file}")

    if args.analyze or args.collect_and_analyze:
        wb = create_analysis_sheets(wb)
        wb.save(workbook_file)
        LOGGER.info(f"Created analysis sheets for books in {workbook_file.name}")

        wb = create_summary_sheet(wb)
        wb = create_correlation_sheet(wb)
        wb = add_charts_to_workbook(wb)
        wb.save(workbook_file)
        LOGGER.info(f"Updated summary sheet in {workbook_file.name}")
        LOGGER.info(f"Created correlations sheet in {workbook_file}")
        LOGGER.info(f"Added chart sheets to {workbook_file.name}")

        plot_diverging_deltas(wb, output_folder=main_folder)

    return 0


if __name__ == "__main__":
    main()
