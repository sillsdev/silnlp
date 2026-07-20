import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

from ..common.environment import SilNlpEnv

# Canonical column set and order for the combined output. Series/Experiment/Steps
# are synthesized by this script; the rest come from the input scores files.
CURRENT_STYLE_COLUMNS = [
    "Series",
    "Experiment",
    "Steps",
    "book",
    "draft_index",
    "src_iso",
    "trg_iso",
    "num_refs",
    "references",
    "sent_len",
    "BLEU",
    "BLEU_1gram_prec",
    "BLEU_2gram_prec",
    "BLEU_3gram_prec",
    "BLEU_4gram_prec",
    "BLEU_brevity_penalty",
    "BLEU_total_sys_len",
    "BLEU_total_ref_len",
    "chrF3",
    "chrF3+",
    "chrF3++",
    "spBLEU",
    "confidence",
]

# Columns to hide in Excel output
COLUMNS_TO_HIDE = [
    "BLEU_1gram_prec",
    "BLEU_2gram_prec",
    "BLEU_3gram_prec",
    "BLEU_4gram_prec",
    "BLEU_brevity_penalty",
    "BLEU_total_sys_len",
    "BLEU_total_ref_len",
    "chrF3",
    "chrF3+",
    "book",
    "draft_index",
    "num_refs",
    "references",
    "sent_len",
    "spBLEU",
    "confidence",
]

# Columns that should be treated as numbers (for coercion and sorting).
NUMERIC_COLUMNS = [
    "BLEU",
    "BLEU_1gram_prec",
    "BLEU_2gram_prec",
    "BLEU_3gram_prec",
    "BLEU_4gram_prec",
    "BLEU_brevity_penalty",
    "BLEU_total_sys_len",
    "BLEU_total_ref_len",
    "chrF3",
    "chrF3+",
    "chrF3++",
    "spBLEU",
    "confidence",
    "num_refs",
    "sent_len",
    "draft_index",
]

# Columns that should be treated as free text.
STRING_COLUMNS = ["Series", "Experiment", "Steps", "book", "src_iso", "trg_iso", "references"]


def check_for_lock_file(folder: Path, filename: str, file_type: str) -> None:
    """Check for lock files and ask the user to close them then exit."""

    if file_type[0] == ".":
        file_type = file_type[1:]

    if file_type.lower() == "csv":
        lockfile = folder / f".~lock.{filename}.{file_type}#"
    elif file_type.lower() == "xlsx":
        lockfile = folder / f"~${filename}.{file_type}"
    else:
        raise ValueError(f"Unsupported file_type for lock file check: {file_type}")

    if lockfile.is_file():
        print(f"Found lock file: {lockfile}")
        print(f"Please close {filename}.{file_type} in folder {folder} OR delete the lock file and try again.")
        sys.exit()


def load_scores(folder: Path) -> pd.DataFrame:
    """Read every folder/<series>/<experiment>/scores-*.csv file into one DataFrame.

    Scores files live in per-experiment subfolders (folder/<experiment>/scores-*.csv),
    and may be nested more deeply, so recurse. parts[-2]/parts[-3] are the file's
    immediate parent (experiment) and grandparent (series) directories.
    ``skipinitialspace`` drops the leading blanks in the comma-space separated files,
    and reading everything as strings preserves the raw values for later coercion.
    """
    csv_files = sorted(folder.rglob("*/scores-*.csv"))
    if not csv_files:
        print(f"No scores csv files were found in folder {folder.resolve()}")
        sys.exit(0)

    frames = []
    for csv_file in csv_files:
        print(f"Processing {csv_file}")
        try:
            df = pd.read_csv(csv_file, dtype=str, skipinitialspace=True, skip_blank_lines=True, on_bad_lines="warn")
        except pd.errors.EmptyDataError:
            print(f"Skipping empty file {csv_file}")
            continue

        # Strip any whitespace left around header names so column lookups are reliable.
        df.columns = df.columns.astype(str).str.strip()
        df.insert(0, "Series", csv_file.parts[-3])
        df.insert(1, "Experiment", csv_file.parts[-2])
        df.insert(2, "Steps", csv_file.stem.split("-")[-1])
        frames.append(df)

    if not frames:
        # Return an empty DataFrame with expected columns so downstream code is safe.
        return pd.DataFrame(columns=CURRENT_STYLE_COLUMNS)

    combined = pd.concat(frames, ignore_index=True, sort=False)
    # Normalize to the canonical column set/order; missing columns become NaN.
    return combined.reindex(columns=CURRENT_STYLE_COLUMNS)


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned_df = df.copy()

    # Strip string columns BEFORE any value comparison so whitespace-padded values
    # (e.g. " ALL") are matched correctly by the filter below.
    for col in STRING_COLUMNS:
        if col in cleaned_df.columns:
            cleaned_df[col] = cleaned_df[col].astype(str).str.strip()

    # Rows with 'ALL' in trg_iso and no chrF3++ score are not useful, so drop them.
    if "trg_iso" in cleaned_df.columns and "chrF3++" in cleaned_df.columns:
        chrf3pp = cleaned_df["chrF3++"]
        missing_chrf3pp = chrf3pp.isna() | (chrf3pp.astype(str).str.strip() == "")
        cleaned_df = cleaned_df[~((cleaned_df["trg_iso"] == "ALL") & missing_chrf3pp)]

    for col in NUMERIC_COLUMNS:
        if col in cleaned_df.columns:
            cleaned_df[col] = pd.to_numeric(cleaned_df[col], errors="coerce")

    return cleaned_df


def sort_dataframe(df: pd.DataFrame, sort_by) -> pd.DataFrame:
    sort_cols = [col for col, _ in sort_by]
    sort_ascending = [asc for _, asc in sort_by]
    return df.sort_values(by=sort_cols, ascending=sort_ascending, na_position="last")


def write_to_excel(df: pd.DataFrame, folder: Path, output_filename: str) -> None:
    output_file = folder / f"{output_filename}.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Scores", index=False)
    wb = openpyxl.load_workbook(output_file)
    ws = wb["Scores"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, col_name in enumerate(header_row, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_name in COLUMNS_TO_HIDE:
            ws.column_dimensions[col_letter].hidden = True
        else:
            max_length = len(str(col_name)) if col_name else 0
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(output_file)
    print(f"Wrote scores to {output_file}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Aggregate CSV files in a folder.")
    parser.add_argument("folder", type=Path, help="Path to the folder containing CSV files.")
    parser.add_argument(
        "--output_filename",
        type=str,
        default="scores",
        help="Filename suffix without the '.csv' or '.xlsx'. \
            The folder name is added as a prefix to make it easier to distinguish scores files in search results.",
    )
    args = parser.parse_args()

    environment = SilNlpEnv.create_standard_environment()
    folder = Path(args.folder)
    base_filename = f"{folder.name}_{args.output_filename}"

    if not folder.is_dir():
        folder = Path(environment.mt_experiments_dir) / args.folder

    # Check for lock files and ask the user to close them.
    check_for_lock_file(folder, base_filename, "xlsx")

    # Read, clean and sort the scores.
    combined_df = load_scores(folder)
    clean_df = clean_dataframe(combined_df)
    sorted_df = sort_dataframe(clean_df, sort_by=[("Series", True), ("chrF3++", False), ("BLEU", False)])

    # Write the data to an excel file
    write_to_excel(sorted_df, folder, base_filename)


if __name__ == "__main__":
    main()
